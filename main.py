import os
import sys
import logging
import requests
import re
import hashlib
import time
from io import BytesIO
from flask import Flask, request
import PyPDF2
from pdf2image import convert_from_bytes
import pytesseract
from PIL import Image
import sqlite3
import json
from datetime import datetime, timedelta
import openpyxl
import numpy as np
import cv2

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

BOT_TOKEN = os.getenv("BOT_TOKEN")
WEBHOOK_URL = os.getenv("WEBHOOK_URL")
DB_PATH = os.getenv("DB_PATH", "bot.db")
ADMIN_CHAT_ID = os.getenv("ADMIN_CHAT_ID", "364191893")

# Режимы экономии ресурсов (для Render free tier)
LOW_RESOURCE = os.getenv("LOW_RESOURCE", "1") == "1"
OCR_DPI = int(os.getenv("OCR_DPI", "150"))  # ниже DPI — меньше память и CPU
OCR_MAX_WORKERS = int(os.getenv("OCR_MAX_WORKERS", "1"))  # 1 поток по умолчанию
OCR_TEXT_THRESHOLD_CHARS = int(os.getenv("OCR_TEXT_THRESHOLD_CHARS", "30"))  # порог длины текста, чтобы считать страницу текстовой
OCR_ENABLE_RETRY = os.getenv("OCR_ENABLE_RETRY", "1") == "1"  # включить повторную попытку для сложных страниц
OCR_RETRY_SCALE = float(os.getenv("OCR_RETRY_SCALE", "1.8"))  # масштаб апскейла при повторе
OCR_RETRY_EXTRA_PSMS = os.getenv("OCR_RETRY_EXTRA_PSMS", "1,11,12,13")  # дополнительные PSM для сложных макетов
OCR_FUTURE_TIMEOUT_SEC = int(os.getenv("OCR_FUTURE_TIMEOUT_SEC", "60"))  # таймаут на распознавание одной страницы
OCR_TOTAL_TIMEOUT_SEC = int(os.getenv("OCR_TOTAL_TIMEOUT_SEC", "180"))  # общий лимит времени на документ
OCR_RETRY_MAX_PSMS = int(os.getenv("OCR_RETRY_MAX_PSMS", "2"))  # ограничить число PSM при ретрае
OCR_MAX_UPSCALED_DIM = int(os.getenv("OCR_MAX_UPSCALED_DIM", "2000"))  # максимум пикселей по стороне после апскейла

if not BOT_TOKEN or not WEBHOOK_URL:
    logger.critical("❌ BOT_TOKEN or WEBHOOK_URL not set!")
    sys.exit(1)

app = Flask(__name__)

# Кэш для отслеживания обработанных сообщений
processed_messages = set()

# Состояния пользователей (ожидают ли они загрузки файла)
user_states = {}

# Временное хранилище для больших OCR-PDF, ожидающих выбора пользователя
pending_files = {}
# Ожидание текстового комментария по conversion_id
awaiting_comment = {}
# Дедупликация callback-запросов, чтобы не перезапускать обработку
processed_callback_ids = set()

# --- БД и аналитика ---
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id TEXT NOT NULL,
                event TEXT NOT NULL,
                meta TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS errors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id TEXT,
                error_code TEXT,
                message TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS feedback (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id TEXT NOT NULL,
                conversion_id TEXT NOT NULL,
                rating INTEGER CHECK (rating BETWEEN 1 AND 5),
                comment TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(user_id, conversion_id)
            )
            """
        )
        conn.commit()
        conn.close()
    except Exception as e:
        logger.exception("💥 Ошибка инициализации БД")

def log_event(user_id, event, meta=None):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO events (user_id, event, meta, created_at) VALUES (?, ?, ?, ?)",
            (str(user_id), event, json.dumps(meta) if meta is not None else None, datetime.utcnow().isoformat())
        )
        conn.commit()
        conn.close()
    except Exception as e:
        logger.exception("💥 Ошибка записи события в БД")

def log_error(user_id, error_code, message):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO errors (user_id, error_code, message, created_at) VALUES (?, ?, ?, ?)",
            (str(user_id) if user_id else None, error_code, str(message), datetime.utcnow().isoformat())
        )
        conn.commit()
        conn.close()
    except Exception as e:
        logger.exception("💥 Ошибка записи ошибки в БД")

def save_feedback(user_id, conversion_id, rating=None, comment=None):
    try:
        conn = get_db()
        cur = conn.cursor()
        # upsert по паре (user_id, conversion_id)
        cur.execute(
            """
            INSERT INTO feedback (user_id, conversion_id, rating, comment, created_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(user_id, conversion_id) DO UPDATE SET
                rating=COALESCE(excluded.rating, feedback.rating),
                comment=COALESCE(excluded.comment, feedback.comment)
            """,
            (str(user_id), str(conversion_id), int(rating) if rating is not None else None, comment, datetime.utcnow().isoformat())
        )
        conn.commit()
        conn.close()
    except Exception as e:
        logger.exception("💥 Ошибка записи оценки в БД")

def get_feedback(user_id, conversion_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "SELECT user_id, conversion_id, rating, comment, created_at FROM feedback WHERE user_id=? AND conversion_id=?",
            (str(user_id), str(conversion_id))
        )
        row = cur.fetchone()
        conn.close()
        return row
    except Exception as e:
        logger.exception("💥 Ошибка чтения feedback из БД")
        return None

def build_rating_keyboard(conversion_id):
    return {
        "inline_keyboard": [[
            {"text": "⭐", "callback_data": f"RATE_1|{conversion_id}"},
            {"text": "⭐⭐", "callback_data": f"RATE_2|{conversion_id}"},
            {"text": "⭐⭐⭐", "callback_data": f"RATE_3|{conversion_id}"},
        ], [
            {"text": "⭐⭐⭐⭐", "callback_data": f"RATE_4|{conversion_id}"},
            {"text": "⭐⭐⭐⭐⭐", "callback_data": f"RATE_5|{conversion_id}"},
        ], [
            {"text": "🗒️ Оставить комментарий", "callback_data": f"FB_COMMENT|{conversion_id}"}
        ]]
    }

# Основное описание бота для показа по кнопке
DESCRIPTION_MESSAGE = (
    "Описание бота (PDF → Текст)\n\n"
    "Бот конвертирует PDF-файлы в обычный текст через OCR и отправляет результат в чат.\n\n"
    "Особенности\n\n"
    "✅  OCR (Машинное зрение): все PDF обрабатываются через OCR для стабильного качества\n"
    "✅  Языки: русский и английский\n"
    "✅  Результат: отправляет .txt с извлечённым текстом\n"
    "✅  Улучшенная обработка:\n\n"
    "- множественные методы предобработки изображений\n"
    "- автоматический выбор лучшего результата OCR\n"
    "- исправление типичных ошибок распознавания\n"
    "- улучшенная очистка текста: склейка переносов, удаление лишних пробелов\n\n"
    "Ограничения\n\n"
    "❗ Только PDF: изображения, DOCX и др. не принимаются\n"
    "❗ Размер файла: максимум — 20 МБ\n"
    "❗ Форматирование: исходное оформление/таблицы/колонки могут потеряться — на выходе чистый текст\n"
    "❗ Качество OCR: зависит от качества скана (разрешение, контраст, шум)\n\n"
    "Как пользоваться\n\n"
    "1⃣   Нажмите кнопку 📤 Отправить PDF на конвертацию\n"
    "2⃣   Отправьте PDF-файл (до 20 МБ)\n"
    "✅  Дождитесь обработки — бот пришлёт .txt с текстом\n\n"
    "Если нужна поддержка больших файлов: сожмите PDF, разбейте по 10 страниц, или отправьте ссылку на файл."
)

def get_main_keyboard():
    return {
        "keyboard": [[
            {"text": "📤 Отправить PDF на конвертацию"},
            {"text": "Возможности и ограничения"}
        ]],
        "resize_keyboard": True,
        "one_time_keyboard": False
    }

def get_message_hash(message):
    """Создает уникальный хеш для сообщения"""
    message_str = str(message.get('message_id', '')) + str(message.get('date', ''))
    return hashlib.md5(message_str.encode()).hexdigest()

def is_message_processed(message_hash):
    """Проверяет, было ли сообщение уже обработано"""
    return message_hash in processed_messages

def mark_message_processed(message_hash):
    """Отмечает сообщение как обработанное"""
    processed_messages.add(message_hash)
    # Очищаем старые записи (старше 1 часа) если их слишком много
    if len(processed_messages) > 1000:
        logger.info("🧹 Очищаем кэш обработанных сообщений")
        processed_messages.clear()

def set_user_waiting_for_file(chat_id, waiting=True):
    """Устанавливает состояние ожидания файла для пользователя"""
    if waiting:
        user_states[chat_id] = {"waiting_for_file": True, "timestamp": time.time()}
    else:
        user_states.pop(chat_id, None)

def is_user_waiting_for_file(chat_id):
    """Проверяет, ожидает ли пользователь загрузки файла"""
    if chat_id not in user_states:
        return False
    
    # Проверяем, не истекло ли время ожидания (30 минут)
    if time.time() - user_states[chat_id]["timestamp"] > 1800:
        user_states.pop(chat_id, None)
        return False
    
    return user_states[chat_id].get("waiting_for_file", False)

def send_message(chat_id, text, reply_markup=None):
    """Отправляет сообщение в Telegram с улучшенной обработкой ошибок"""
    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        data = {"chat_id": chat_id, "text": text}
        if reply_markup:
            data["reply_markup"] = reply_markup
        
        response = requests.post(url, json=data, timeout=10)
        if not response.ok:
            logger.error(f"❌ Ошибка отправки сообщения: {response.status_code} - {response.text}")
        else:
            logger.info(f"✅ Сообщение отправлено в чат {chat_id}")
    except requests.exceptions.Timeout:
        logger.error("⏰ Таймаут при отправке сообщения")
    except Exception as e:
        logger.exception("💥 Ошибка при отправке сообщения")

def send_document(chat_id, file_buffer, filename):
    """Отправляет документ в Telegram с улучшенной обработкой ошибок"""
    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
        files = {"document": (filename, file_buffer, "text/plain")}
        data = {"chat_id": chat_id}
        
        response = requests.post(url, files=files, data=data, timeout=60)
        if not response.ok:
            logger.error(f"❌ Ошибка отправки документа: {response.status_code} - {response.text}")
        else:
            logger.info(f"✅ Документ {filename} отправлен в чат {chat_id}")
    except requests.exceptions.Timeout:
        logger.error("⏰ Таймаут при отправке документа")
    except Exception as e:
        logger.exception("💥 Ошибка при отправке документа")

def send_binary_document(chat_id, file_buffer, filename, mime_type):
    """Отправляет бинарный документ (например, .xlsx)."""
    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
        files = {"document": (filename, file_buffer, mime_type)}
        data = {"chat_id": chat_id}
        response = requests.post(url, files=files, data=data, timeout=60)
        if not response.ok:
            logger.error(f"❌ Ошибка отправки бинарного документа: {response.status_code} - {response.text}")
        else:
            logger.info(f"✅ Документ {filename} отправлен в чат {chat_id}")
    except requests.exceptions.Timeout:
        logger.error("⏰ Таймаут при отправке бинарного документа")
    except Exception as e:
        logger.exception("💥 Ошибка при отправке бинарного документа")

def generate_excel_stats(last_days=30):
    """Генерирует Excel со статистикой за последние N дней."""
    cutoff = (datetime.utcnow() - timedelta(days=last_days)).isoformat()
    conn = get_db()
    cur = conn.cursor()
    # Overview
    cur.execute("SELECT COUNT(DISTINCT user_id) FROM events WHERE created_at >= ?", (cutoff,))
    unique_users = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM events WHERE event='file_received' AND created_at >= ?", (cutoff,))
    file_uses = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM events WHERE event='ocr_success' AND created_at >= ?", (cutoff,))
    ocr_success = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM events WHERE event='ocr_error' AND created_at >= ?", (cutoff,))
    ocr_errors = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM errors WHERE created_at >= ?", (cutoff,))
    total_errors = cur.fetchone()[0] or 0
    cur.execute("SELECT rating, COUNT(*) FROM feedback WHERE rating IS NOT NULL AND created_at >= ? GROUP BY rating ORDER BY rating", (cutoff,))
    ratings_rows = cur.fetchall()
    # Daily events
    cur.execute(
        """
        SELECT substr(created_at,1,10) as day, event, COUNT(*) c
        FROM events
        WHERE created_at >= ?
        GROUP BY day, event
        ORDER BY day, event
        """ , (cutoff,)
    )
    daily_events = cur.fetchall()
    # Errors detail
    cur.execute(
        """
        SELECT substr(created_at,1,10) as day, error_code, COUNT(*) c
        FROM errors
        WHERE created_at >= ?
        GROUP BY day, error_code
        ORDER BY c DESC
        """, (cutoff,)
    )
    errors_agg = cur.fetchall()
    cur.execute("SELECT created_at, user_id, error_code, message FROM errors WHERE created_at >= ? ORDER BY created_at DESC LIMIT 1000", (cutoff,))
    errors_raw = cur.fetchall()
    # Feedback raw (включая комментарии и conversion_id)
    cur.execute("SELECT created_at, user_id, conversion_id, rating, comment FROM feedback WHERE created_at >= ? ORDER BY created_at DESC", (cutoff,))
    feedback_raw = cur.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws_overview = wb.active
    ws_overview.title = "Overview"
    ws_overview.append(["Metric", "Value"])
    ws_overview.append(["Unique users", unique_users])
    ws_overview.append(["File uses", file_uses])
    ws_overview.append(["OCR success", ocr_success])
    ws_overview.append(["OCR errors", ocr_errors])
    ws_overview.append(["Errors total", total_errors])
    ws_overview.append(["Ratings (rating:count)", ", ".join([f"{r[0]}:{r[1]}" for r in ratings_rows]) if ratings_rows else "-"])

    ws_events = wb.create_sheet("DailyEvents")
    ws_events.append(["Day", "Event", "Count"])
    for row in daily_events:
        ws_events.append(list(row))

    ws_errors = wb.create_sheet("ErrorsAgg")
    ws_errors.append(["Day", "ErrorCode", "Count"])
    for row in errors_agg:
        ws_errors.append(list(row))

    ws_errors_raw = wb.create_sheet("ErrorsRaw")
    ws_errors_raw.append(["CreatedAt", "UserId", "ErrorCode", "Message"])
    for row in errors_raw:
        ws_errors_raw.append(list(row))

    ws_feedback = wb.create_sheet("Feedback")
    ws_feedback.append(["CreatedAt", "UserId", "ConversionId", "Rating", "Comment"])
    for row in feedback_raw:
        ws_feedback.append(list(row))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def answer_callback_query(callback_query_id, text=None):
    """Отвечает на callback-запрос для снятия индикатора загрузки на кнопке"""
    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/answerCallbackQuery"
        data = {"callback_query_id": callback_query_id}
        if text:
            data["text"] = text
        response = requests.post(url, data=data, timeout=10)
        if not response.ok:
            logger.error(f"❌ Ошибка answerCallbackQuery: {response.status_code} - {response.text}")
    except Exception as e:
        logger.exception("💥 Ошибка при answerCallbackQuery")

def build_split_options_keyboard():
    """Создает inline-клавиатуру с вариантами обработки большого OCR-PDF"""
    return {
        "inline_keyboard": [[
            {"text": "✂️ разделить файл", "callback_data": "SPLIT_PDF"},
            {"text": "🔟 распознать только первые 10 страниц", "callback_data": "OCR_FIRST_10"}
        ]]
    }

def clean_text(text):
    """Очищает извлеченный текст с улучшенной постобработкой"""
    if not text:
        return ""
    
    # Исправляем переносы слов
    text = re.sub(r'([а-яА-Яa-zA-Z])-\n([а-яА-Яa-zA-Z])', r'\1\2', text)
    
    # Заменяем одиночные переносы на пробелы
    text = re.sub(r'(?<!\n)\n(?!\n)', ' ', text)
    
    # Убираем лишние пробелы
    text = re.sub(r' +', ' ', text)
    
    # Исправляем слипшиеся слова (добавляем пробелы между заглавными буквами)
    text = re.sub(r'([а-я])([А-Я])', r'\1 \2', text)
    text = re.sub(r'([a-z])([A-Z])', r'\1 \2', text)
    
    # Исправляем слипшиеся цифры и буквы
    text = re.sub(r'([а-яА-Яa-zA-Z])(\d)', r'\1 \2', text)
    text = re.sub(r'(\d)([а-яА-Яa-zA-Z])', r'\1 \2', text)
    
    # Исправляем слипшиеся знаки препинания
    text = re.sub(r'([а-яА-Яa-zA-Z])([.,:;!?])', r'\1\2', text)
    text = re.sub(r'([.,:;!?])([а-яА-Яa-zA-Z])', r'\1 \2', text)
    
    # Исправляем типичные OCR ошибки
    ocr_fixes = {
        '0': 'О',  # Цифра 0 вместо буквы О в начале слов
        '1': 'I',  # Цифра 1 вместо буквы I
        '5': 'S',  # Цифра 5 вместо буквы S
        '8': 'B',  # Цифра 8 вместо буквы B
    }
    
    # Применяем исправления только в контексте слов
    for wrong, correct in ocr_fixes.items():
        # Заменяем цифры на буквы в словах (но не в числах)
        text = re.sub(rf'\b{wrong}([а-яА-Яa-zA-Z])', rf'{correct}\1', text)
        text = re.sub(rf'([а-яА-Яa-zA-Z]){wrong}\b', rf'\1{correct}', text)
    
    # Убираем лишние пробелы после исправлений
    text = re.sub(r' +', ' ', text)
    
    # Форматируем строки
    text = '\n'.join(line.strip() for line in text.splitlines())
    
    return text.strip()

def pil_to_cv(img_pil):
    arr = np.array(img_pil)
    if arr.ndim == 2:
        return arr
    # PIL is RGB, cv2 expects BGR
    return cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)

def cv_to_pil(img_cv):
    if len(img_cv.shape) == 2:
        rgb = cv2.cvtColor(img_cv, cv2.COLOR_GRAY2RGB)
    else:
        rgb = cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB)
    return Image.fromarray(rgb)

def deskew_image(gray):
    # Estimate skew angle and rotate to deskew
    try:
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
        inv = 255 - thresh
        coords = np.column_stack(np.where(inv > 0))
        angle = 0.0
        if coords.size > 0:
            rect = cv2.minAreaRect(coords)
            angle = rect[-1]
            if angle < -45:
                angle = -(90 + angle)
            else:
                angle = -angle
        (h, w) = gray.shape[:2]
        center = (w // 2, h // 2)
        M = cv2.getRotationMatrix2D(center, angle, 1.0)
        rotated = cv2.warpAffine(gray, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
        return rotated
    except Exception:
        return gray

def preprocess_image_for_ocr(img_pil):
    """Агрессивная предобработка для плохих сканов с множественными методами."""
    img_cv = pil_to_cv(img_pil)
    if img_cv.ndim == 3:
        gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
    else:
        gray = img_cv
    
    # Оптимизированное масштабирование для Render (ограниченная память)
    scale_factor = 2.0  # Уменьшаем масштаб для экономии памяти
    if gray.shape[0] < 1500 or gray.shape[1] < 1500:
        new_width = int(gray.shape[1] * scale_factor)
        new_height = int(gray.shape[0] * scale_factor)
        gray = cv2.resize(gray, (new_width, new_height), interpolation=cv2.INTER_CUBIC)
    
    # Агрессивное улучшение контраста
    clahe = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8,8))
    gray = clahe.apply(gray)
    
    # Дополнительное выравнивание гистограммы
    gray = cv2.equalizeHist(gray)
    
    # Улучшенное шумоподавление
    gray = cv2.bilateralFilter(gray, d=11, sigmaColor=100, sigmaSpace=100)
    gray = cv2.medianBlur(gray, 5)
    
    # Множественные методы бинаризации и выбор лучшего
    methods = []
    
    # Метод 1: Адаптивная пороговая обработка
    thr1 = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                cv2.THRESH_BINARY, 25, 15)
    methods.append(thr1)
    
    # Метод 2: Otsu thresholding
    _, thr2 = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    methods.append(thr2)
    
    # Метод 3: Адаптивная с другим ядром
    thr3 = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C,
                                cv2.THRESH_BINARY, 21, 12)
    methods.append(thr3)
    
    # Выбираем метод с лучшим соотношением белого/черного
    best_thr = None
    best_ratio = 0
    for thr in methods:
        white_pixels = np.sum(thr == 255)
        total_pixels = thr.shape[0] * thr.shape[1]
        ratio = white_pixels / total_pixels
        if 0.1 < ratio < 0.9:  # Ищем сбалансированное изображение
            if abs(ratio - 0.5) < abs(best_ratio - 0.5):
                best_ratio = ratio
                best_thr = thr
    
    if best_thr is None:
        best_thr = methods[0]  # Fallback на первый метод
    
    # Агрессивная морфологическая очистка
    # Удаляем мелкий шум
    kernel_noise = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    cleaned = cv2.morphologyEx(best_thr, cv2.MORPH_OPEN, kernel_noise, iterations=2)
    
    # Закрываем разрывы в символах
    kernel_close = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 1))
    cleaned = cv2.morphologyEx(cleaned, cv2.MORPH_CLOSE, kernel_close, iterations=1)
    
    # Удаляем очень мелкие объекты (шум)
    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(255 - cleaned, connectivity=8)
    for i in range(1, num_labels):
        if stats[i, cv2.CC_STAT_AREA] < 50:  # Удаляем объекты меньше 50 пикселей
            cleaned[labels == i] = 255
    
    # Исправление наклона
    cleaned = deskew_image(cleaned)
    
    return cv_to_pil(cleaned)

def enhance_ocr_with_alternatives(img_pil):
    """Дополнительные методы улучшения OCR для сложных случаев"""
    try:
        # В low-resource режиме ограничиваем варианты для экономии CPU/RAM
        if LOW_RESOURCE:
            return [("original", preprocess_image_for_ocr(img_pil))]

        variants = []
        variants.append(("original", preprocess_image_for_ocr(img_pil)))

        img_cv = pil_to_cv(img_pil)
        if img_cv.ndim == 3:
            gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
        else:
            gray = img_cv
        inverted = 255 - gray
        variants.append(("inverted", cv_to_pil(inverted)))
        high_contrast = cv2.convertScaleAbs(gray, alpha=2.0, beta=0)
        variants.append(("high_contrast", cv_to_pil(high_contrast)))
        blurred = cv2.GaussianBlur(gray, (3, 3), 0)
        variants.append(("blurred", cv_to_pil(blurred)))
        return variants
    except Exception as e:
        logger.warning(f"⚠️ Ошибка создания вариантов изображения: {e}")
        return [("original", img_pil)]

def handle_file_questions(text):
    """Обрабатывает вопросы о файлах и ограничениях"""
    text_lower = text.lower()
    
    # Ключевые слова для вопросов о файлах
    file_keywords = ['файл', 'файлы', 'отправ', 'загру', 'pdf', 'формат', 'тип', 'размер', 'ограничен']
    question_keywords = ['какие', 'что', 'можно', 'как', 'какой', 'сколько', 'максимальн', 'лимит']
    
    is_file_question = any(keyword in text_lower for keyword in file_keywords)
    is_question = any(keyword in text_lower for keyword in question_keywords)
    
    if is_file_question and is_question:
        return """📋 **Информация о файлах:**

✅ **Принимаемые форматы:** Только PDF файлы
📏 **Максимальный размер:** 20 МБ
⏱️ **Время обработки:** 1-4 минуты (все PDF через OCR)
📄 **Ограничения:** 
   • Максимум 10 страниц для больших файлов
   • Поддержка русского и английского языков
   • Все PDF обрабатываются через OCR для лучшего качества

💡 **Как отправить файл:**
1. Нажмите кнопку "📤 Отправить PDF на конвертацию"
2. Выберите PDF файл
3. Дождитесь обработки

❓ **Нужна помощь?** Используйте /start для перезапуска"""
    
    return None

def extract_text_from_pdf(file_bytes, is_ocr_needed=False, progress_callback=None, first_page=None, last_page=None, max_pages_default=10):
    """Гибридное извлечение текста: нативный текст из PDF, OCR только для изображений.

    - В low-resource режиме используем по-страничную обработку и сниженный DPI.
    - Если страница содержит нативный текст, берем его без OCR.
    - Если страница пустая или без текста — включаем OCR только для этой страницы.
    - Можно ограничить диапазон страниц (first_page/last_page) и общий лимит страниц.
    """
    try:
        fp = first_page if first_page is not None else 1
        lp = last_page if last_page is not None else max_pages_default
        if lp < fp:
            fp, lp = lp, fp

        try:
            reader = PyPDF2.PdfReader(BytesIO(file_bytes))
            actual_pages = len(reader.pages)
            lp = min(lp, actual_pages)
        except Exception as e:
            logger.warning(f"⚠️ Не удалось определить количество страниц: {e}")
            reader = None
            actual_pages = lp - fp + 1

        total_pages = lp - fp + 1
        logger.info(f"📄 Обработка страниц {fp}-{lp} (всего {total_pages})")

        collected_text = []
        start_time = time.time()

        # Последовательная обработка страниц для экономии памяти
        for page_index in range(fp - 1, lp):
            page_num_human = page_index + 1
            try:
                # Проверка общего таймаута на документ
                if OCR_TOTAL_TIMEOUT_SEC and (time.time() - start_time) > OCR_TOTAL_TIMEOUT_SEC:
                    logger.warning("⏰ Истек общий таймаут обработки документа, возвращаем частичный результат")
                    break
                page_text = ""
                if reader is not None:
                    try:
                        page = reader.pages[page_index]
                        # В PyPDF2 3.x: extract_text() может вернуть None
                        native_text = page.extract_text() or ""
                        # Отбрасываем шум: слишком короткие строки — вероятно артефакты
                        native_text_clean = native_text.strip()
                        if len(native_text_clean) >= OCR_TEXT_THRESHOLD_CHARS:
                            page_text = native_text_clean
                            logger.info(f"📝 p.{page_num_human}: использован нативный текст ({len(page_text)} симв.)")
                    except Exception as e:
                        logger.debug(f"⚠️ p.{page_num_human}: ошибка нативного извлечения: {e}")

                if not page_text:
                    # OCR только этой страницы
                    dpi = OCR_DPI if LOW_RESOURCE else 200
                    try:
                        images = convert_from_bytes(
                            file_bytes,
                            dpi=dpi,
                            first_page=page_num_human,
                            last_page=page_num_human
                        )
                        if images:
                            page_text = process_image_chunk(images, progress_callback)
                            logger.info(f"🔎 p.{page_num_human}: OCR ({dpi} DPI) -> {len(page_text)} симв.")
                        else:
                            logger.warning(f"⚠️ p.{page_num_human}: не удалось конвертировать в изображение")
                    except Exception as e:
                        logger.error(f"❌ p.{page_num_human}: ошибка конвертации/OCR: {e}")

                collected_text.append(page_text)
                if progress_callback:
                    progress_callback(f"✅ Страница {page_num_human} завершена")
            except Exception as e:
                logger.error(f"❌ p.{page_num_human}: общая ошибка обработки страницы: {e}")
                collected_text.append("")

        return clean_text("\n".join(collected_text))
    except Exception:
        logger.exception("💥 Ошибка гибридного извлечения")
        raise

def process_image_chunk(images, progress_callback=None):
    """Обрабатывает часть изображений для OCR"""
    if not images:
        logger.warning("⚠️ Нет изображений для обработки")
        return ""
    
    ocr_text = ""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    def ocr_single(idx_img):
        i, img = idx_img
        try:
            # Оптимизируем размер изображения для Render (ограниченная память)
            if img.width > 2000 or img.height > 2000:
                img.thumbnail((2000, 2000), Image.Resampling.LANCZOS)
            
            # Расширенный whitelist для лучшего распознавания
            safe_whitelist = (
                "0123456789"
                "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
                "abcdefghijklmnopqrstuvwxyz"
                "АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ"
                "абвгдеёжзийклмнопрстуфхцчшщъыьэюя"
                ".,:;!?()\-–—_№%$€₽«»""''"
            )
            
            # Множественные конфигурации (урезаем в low-resource)
            if LOW_RESOURCE:
                configs = [
                    f"--psm 6 --oem 3 -c tessedit_char_whitelist={safe_whitelist} -c preserve_interword_spaces=1",
                    f"--psm 4 --oem 3 -c tessedit_char_whitelist={safe_whitelist} -c preserve_interword_spaces=1",
                ]
            else:
                configs = [
                    f"--psm 6 --oem 3 -c tessedit_char_whitelist={safe_whitelist} -c preserve_interword_spaces=1",
                    f"--psm 4 --oem 3 -c tessedit_char_whitelist={safe_whitelist} -c preserve_interword_spaces=1",
                    f"--psm 3 --oem 3 -c tessedit_char_whitelist={safe_whitelist} -c preserve_interword_spaces=1",
                    f"--psm 8 --oem 3 -c tessedit_char_whitelist={safe_whitelist}",
                    "--psm 6 --oem 3 -c preserve_interword_spaces=1",
                    "--psm 4 --oem 3 -c preserve_interword_spaces=1"
                ]
            
            # Получаем варианты изображения для OCR
            img_variants = enhance_ocr_with_alternatives(img)
            
            text = ""
            best_text = ""
            max_length = 0
            
            # Пробуем комбинации вариантов изображения и конфигураций
            for variant_name, variant_img in img_variants:
                for config in configs:
                    try:
                        result = pytesseract.image_to_string(
                            variant_img,
                            lang='rus+eng',
                            config=config
                        )
                        
                        if result.strip():
                            # Выбираем результат с наибольшим количеством символов
                            if len(result.strip()) > max_length:
                                max_length = len(result.strip())
                                best_text = result
                            
                            # Если результат достаточно хорош, используем его
                            min_ok = 80 if LOW_RESOURCE else 100
                            if len(result.strip()) > min_ok:
                                text = result
                                logger.info(f"✅ Найден хороший результат с {variant_name} и конфигом {config[:20]}...")
                                break
                                
                    except Exception as e:
                        logger.warning(f"⚠️ Ошибка OCR с {variant_name} и конфигом {config[:20]}: {e}")
                        continue
                
                if text.strip():  # Если нашли хороший результат, прерываем
                    break
            
            # Если не нашли хороший результат, используем лучший из всех
            if not text.strip() and best_text.strip():
                text = best_text
                logger.info(f"✅ Используем лучший результат из всех попыток ({max_length} символов)")

            # Качественный ретрай: если текст слишком короткий/плохой — апскейл и расширенные PSM
            if OCR_ENABLE_RETRY and len((text or "").strip()) < (60 if LOW_RESOURCE else 90):
                try:
                    retry_variants = []
                    # Берем лучший из существующих как основу, либо оригинал
                    base_img = img_variants[0][1] if img_variants else img
                    # Апскейл с ограничением максимального размера
                    width, height = base_img.size
                    scaled_w = int(width * OCR_RETRY_SCALE)
                    scaled_h = int(height * OCR_RETRY_SCALE)
                    if max(scaled_w, scaled_h) > OCR_MAX_UPSCALED_DIM:
                        scale_clip = OCR_MAX_UPSCALED_DIM / max(scaled_w, scaled_h)
                        scaled_w = max(1, int(scaled_w * scale_clip))
                        scaled_h = max(1, int(scaled_h * scale_clip))
                    new_size = (scaled_w, scaled_h)
                    upscaled = base_img.resize(new_size, Image.Resampling.LANCZOS)
                    retry_variants.append(("upscaled", upscaled))

                    # Дополнительные PSMы для сложных макетов
                    extra_psms = [p.strip() for p in OCR_RETRY_EXTRA_PSMS.split(',') if p.strip()][:OCR_RETRY_MAX_PSMS]
                    # Базовые конфиги с whitelists
                    retry_configs = []
                    for psm in extra_psms:
                        retry_configs.append(
                            f"--psm {psm} --oem 3 -c preserve_interword_spaces=1"
                        )
                        retry_configs.append(
                            f"--psm {psm} --oem 3 -c tessedit_char_whitelist={safe_whitelist} -c preserve_interword_spaces=1"
                        )

                    best_retry = text
                    best_retry_len = len((text or "").strip())
                    for rname, rimg in retry_variants:
                        for rconfig in retry_configs:
                            try:
                                r = pytesseract.image_to_string(rimg, lang='rus+eng', config=rconfig)
                                rlen = len((r or "").strip())
                                if rlen > best_retry_len:
                                    best_retry = r
                                    best_retry_len = rlen
                                    logger.info(f"🔁 Улучшение после ретрая {rname} psm={rconfig.split()[1]} → {rlen} симв.")
                                # Достаточно хороший текст — выходим
                                if rlen >= (120 if LOW_RESOURCE else 160):
                                    break
                            except Exception as e:
                                logger.debug(f"⚠️ Ошибка ретрая OCR: {e}")
                        if best_retry_len >= (120 if LOW_RESOURCE else 160):
                            break
                    if best_retry_len > len((text or "").strip()):
                        text = best_retry
                        logger.info("✅ Ретрай дал лучший результат, используем его")
                except Exception as e:
                    logger.debug(f"⚠️ Ретрай не выполнен: {e}")
            
        except Exception as e:
            logger.error(f"❌ Ошибка OCR на странице {i+1}: {e}")
            text = ""
        return i, text

    # Уменьшаем количество потоков (env-переключатель)
    max_workers = max(1, min(OCR_MAX_WORKERS, len(images)))
    # Если одна страница — без пула потоков, чтобы избежать накладных расходов
    if len(images) == 1:
        i, text = ocr_single((0, images[0]))
        if progress_callback:
            progress_callback(f"✅ Страница {i+1} завершена")
        else:
            logger.info(f"✅ Страница {i+1} завершена")
        ocr_text += text + "\n"
        return clean_text(ocr_text)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(ocr_single, (i, img)): i for i, img in enumerate(images)}
        
        # Добавляем таймаут для предотвращения зависания
        from concurrent.futures import TimeoutError
        for fut in as_completed(futures, timeout=OCR_FUTURE_TIMEOUT_SEC * max(1, len(images))):
            try:
                i, text = fut.result(timeout=OCR_FUTURE_TIMEOUT_SEC)
                if progress_callback:
                    progress_callback(f"✅ Страница {i+1} завершена")
                else:
                    logger.info(f"✅ Страница {i+1} завершена")
                ocr_text += text + "\n"
            except TimeoutError:
                logger.error("⏰ Таймаут обработки страницы")
                continue
            except Exception as e:
                logger.error(f"❌ Ошибка обработки страницы: {e}")
                continue
    logger.info("✅ OCR завершен успешно")
    return clean_text(ocr_text)

@app.route("/webhook", methods=["POST"])
def telegram_webhook():
    """Обрабатывает webhook от Telegram с защитой от дублирования"""
    try:
        data = request.get_json()
        logger.info(f"📨 Получен webhook: {data}")
        
        if not data:
            logger.info("❌ Пустой webhook")
            return "OK", 200

        # Callback-кнопки (inline)
        if "callback_query" in data:
            cb = data["callback_query"]
            callback_id = cb.get("id")
            from_user = cb.get("from", {})
            chat = cb.get("message", {}).get("chat", {})
            chat_id = chat.get("id")
            action = cb.get("data")
            logger.info(f"🖱️ Callback: {action} от {from_user.get('id')} в чате {chat_id}")

            # Дедуп: игнорируем повторные callback'и с тем же ID
            if callback_id in processed_callback_ids:
                logger.info(f"🔁 Пропуск повторного callback_id={callback_id}")
                return "OK", 200

            if callback_id:
                answer_callback_query(callback_id)
                processed_callback_ids.add(callback_id)

            # Обработка оценки качества с conversion_id
            if action and action.startswith("RATE_"):
                try:
                    payload = action.split("_")[1]
                    rating_str, conv_id = payload.split("|")
                    rating = int(rating_str)
                    # Проверяем, не оставлял ли уже пользователь feedback для этой конвертации
                    if get_feedback(chat_id, conv_id):
                        send_message(chat_id, "ℹ️ Оценка/комментарий по этой конвертации уже сохранены.")
                        return "OK", 200
                    if 1 <= rating <= 5:
                        save_feedback(chat_id, conv_id, rating=rating)
                        log_event(chat_id, "feedback", {"rating": rating, "conversion_id": conv_id})
                        send_message(chat_id, "🙏 Спасибо! Ваша оценка помогает нам становиться лучше.")
                    else:
                        send_message(chat_id, "❌ Некорректная оценка.")
                except Exception as e:
                    logger.exception("💥 Ошибка обработки рейтинга")
                    send_message(chat_id, "❌ Не удалось сохранить оценку.")
                return "OK", 200

            # Запрос комментария по conversion_id
            if action and action.startswith("FB_COMMENT|"):
                try:
                    conv_id = action.split("|")[1]
                    # Если уже есть запись feedback для этой пары, не даём повторно
                    if get_feedback(chat_id, conv_id):
                        send_message(chat_id, "ℹ️ Комментарий по этой конвертации уже сохранён.")
                        return "OK", 200
                    awaiting_comment[chat_id] = conv_id
                    send_message(chat_id, "✍️ Напишите, пожалуйста, комментарий: что понравилось/не понравилось/что улучшить.")
                except Exception as e:
                    logger.exception("💥 Ошибка запуска запроса комментария")
                    send_message(chat_id, "❌ Не удалось запросить комментарий.")
                return "OK", 200

            pending = pending_files.get(chat_id)
            if not pending:
                send_message(chat_id, "❌ Не найден файл для обработки. Отправьте PDF заново.")
                return "OK", 200

            file_bytes = pending.get("file_bytes")
            base_name = pending.get("file_name", "converted.pdf")
            total_pages = pending.get("num_pages", 0)

            def progress_callback(msg):
                logger.info(f"📊 {msg}")

            if action == "OCR_FIRST_10":
                send_message(chat_id, "🔟 Начинаю распознавать первые 10 страниц...")
                try:
                    text = extract_text_from_pdf(
                        file_bytes,
                        is_ocr_needed=True,
                        progress_callback=progress_callback,
                        first_page=1,
                        last_page=min(10, total_pages)
                    )
                    if not text.strip():
                        send_message(chat_id, "❌ Не удалось извлечь текст с первых 10 страниц.")
                    else:
                        txt_name = os.path.splitext(base_name)[0] + "_p1-" + str(min(10, total_pages)) + ".txt"
                        txt_buffer = BytesIO(text.encode("utf-8"))
                        send_document(chat_id, txt_buffer, txt_name)
                        send_message(chat_id, "✅ Готово! Отправил результат для первых 10 страниц.")
                except Exception as e:
                    logger.exception("💥 Ошибка при OCR первых 10 страниц")
                    send_message(chat_id, "❌ Произошла ошибка при распознавании первых 10 страниц.")
                finally:
                    # Не очищаем pending_files здесь, т.к. он нужен для SPLIT_PDF цикла
                    set_user_waiting_for_file(chat_id, False)
                return "OK", 200

            if action == "SPLIT_PDF":
                send_message(chat_id, f"✂️ Начинаю делить файл на части по 10 страниц (всего {total_pages}).")
                part_index = 1
                for start in range(1, total_pages + 1, 10):
                    end = min(start + 9, total_pages)
                    send_message(chat_id, f"⏳ Обрабатываю страницы {start}-{end}...")
                    try:
                        part_text = extract_text_from_pdf(
                            file_bytes,
                            is_ocr_needed=True,
                            progress_callback=progress_callback,
                            first_page=start,
                            last_page=end
                        )
                        if not part_text.strip():
                            send_message(chat_id, f"⚠️ Не удалось извлечь текст для страниц {start}-{end}.")
                        else:
                            txt_name = os.path.splitext(base_name)[0] + f"_part{part_index}_p{start}-{end}.txt"
                            txt_buffer = BytesIO(part_text.encode("utf-8"))
                            send_document(chat_id, txt_buffer, txt_name)
                            send_message(chat_id, f"✅ Готово: страницы {start}-{end} отправлены.")
                    except Exception as e:
                        logger.exception(f"💥 Ошибка при обработке страниц {start}-{end}")
                        send_message(chat_id, f"❌ Ошибка при обработке страниц {start}-{end}.")
                    finally:
                        part_index += 1
                send_message(chat_id, "🎉 Все части готовы и отправлены. Можете отправить следующий файл.")
                # Очищаем pending после полного завершения цикла
                pending_files.pop(chat_id, None)
                set_user_waiting_for_file(chat_id, False)
                return "OK", 200

            send_message(chat_id, "❓ Неизвестное действие. Попробуйте снова.")
            return "OK", 200

        if "message" not in data:
            logger.info("❌ Нет сообщения в данных")
            return "OK", 200

        message = data["message"]
        message_id = message.get("message_id", "unknown")
        chat_id = message["chat"]["id"]
        
        # Проверяем, не обрабатывали ли мы это сообщение уже
        message_hash = get_message_hash(message)
        if is_message_processed(message_hash):
            logger.info(f"🔄 Пропускаем дублирующееся сообщение ID: {message_id}")
            return "OK", 200
        
        # Отмечаем сообщение как обрабатываемое
        mark_message_processed(message_hash)
        logger.info(f"📝 Обрабатываем новое сообщение ID: {message_id}")

        if "text" in message:
            text = message["text"]
            # Приоритетно: если ожидаем комментарий, сохраняем его
            if chat_id in awaiting_comment:
                conv_id = awaiting_comment.pop(chat_id)
                try:
                    # Сохраняем комментарий единожды для этой конвертации
                    save_feedback(chat_id, conv_id, comment=text)
                    log_event(chat_id, "feedback_comment", {"conversion_id": conv_id})
                    send_message(chat_id, "✅ Спасибо! Комментарий сохранён.")
                except Exception as e:
                    logger.exception("💥 Ошибка сохранения комментария")
                    send_message(chat_id, "❌ Не удалось сохранить комментарий.")
                return "OK", 200
            if text == "/start":
                # Сбрасываем состояние пользователя
                set_user_waiting_for_file(chat_id, False)
                log_event(chat_id, "start")
                reply_markup = get_main_keyboard()
                send_message(
                    chat_id,
                    "👋 Привет! Я бот для конвертации PDF в текст.\n\nНажмите кнопку ниже, чтобы начать.",
                    reply_markup
                )
            elif text == "/stop":
                set_user_waiting_for_file(chat_id, False)
                send_message(chat_id, "🛑 Бот остановлен. Используйте /start для перезапуска.")
            elif text == "/statistic":
                if ADMIN_CHAT_ID and str(chat_id) == str(ADMIN_CHAT_ID):
                    try:
                        send_message(chat_id, "⏳ Формирую Excel со статистикой за 30 дней...")
                        xlsx_buf = generate_excel_stats(last_days=30)
                        send_binary_document(chat_id, xlsx_buf, "bot_stats_last_30_days.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    except Exception as e:
                        logger.exception("💥 Ошибка формирования Excel статистики")
                        send_message(chat_id, "❌ Не удалось сформировать статистику.")
                else:
                    send_message(chat_id, "⛔ Недостаточно прав.")
            elif text == "/stats":
                if ADMIN_CHAT_ID and str(chat_id) == str(ADMIN_CHAT_ID):
                    try:
                        conn = get_db()
                        cur = conn.cursor()
                        cur.execute("SELECT COUNT(DISTINCT user_id) FROM events")
                        users = cur.fetchone()[0] or 0
                        cur.execute("SELECT COUNT(*) FROM events WHERE event='file_received'")
                        uses = cur.fetchone()[0] or 0
                        cur.execute("SELECT COUNT(*) FROM errors")
                        err_count = cur.fetchone()[0] or 0
                        cur.execute("SELECT error_code, COUNT(*) c FROM errors GROUP BY error_code ORDER BY c DESC LIMIT 5")
                        top_errors = cur.fetchall()
                        cur.execute("SELECT rating, COUNT(*) c FROM feedback GROUP BY rating ORDER BY rating")
                        rating_rows = cur.fetchall()
                        conn.close()
                        ratings = ", ".join([f"{r[0]}★: {r[1]}" for r in rating_rows]) if rating_rows else "нет"
                        top_errs = "\n".join([f"- {r[0]}: {r[1]}" for r in top_errors]) if top_errors else "нет"
                        send_message(
                            chat_id,
                            f"📊 Статистика:\n\n"
                            f"👤 Уникальные пользователи: {users}\n"
                            f"📥 Загрузок файлов: {uses}\n"
                            f"⚠️ Ошибок: {err_count}\n\n"
                            f"Топ ошибок:\n{top_errs}\n\n"
                            f"Оценки: {ratings}"
                        )
                    except Exception as e:
                        logger.exception("💥 Ошибка /stats")
                        send_message(chat_id, "❌ Ошибка при получении статистики.")
                else:
                    send_message(chat_id, "⛔ Недостаточно прав.")
            elif text == "📤 Отправить PDF на конвертацию":
                # Пользователь нажал кнопку - устанавливаем состояние ожидания файла
                set_user_waiting_for_file(chat_id, True)
                log_event(chat_id, "request_upload")
                send_message(
                    chat_id,
                    "📎 Отлично! Теперь отправьте PDF файл для конвертации.\n\n💡 Максимальный размер файла: 20 МБ"
                )
            elif text == "Возможности и ограничения":
                # Показать описание бота
                send_message(chat_id, DESCRIPTION_MESSAGE)
            else:
                # Проверяем, не задает ли пользователь вопрос о файлах
                file_info = handle_file_questions(text)
                if file_info:
                    send_message(chat_id, file_info)
                else:
                    # Если пользователь не в состоянии ожидания файла, склоняем к нажатию кнопки
                    if not is_user_waiting_for_file(chat_id):
                        send_message(
                            chat_id,
                            "📎 Я работаю только с PDF-файлами.\n\n"
                            "💡 **Чтобы отправить файл:**\n"
                            "1. Нажмите кнопку «📤 Отправить PDF на конвертацию»\n"
                            "2. Выберите PDF файл\n"
                            "3. Дождитесь обработки\n\n"
                            "❓ **Вопросы о файлах?** Спросите меня о форматах, размерах или ограничениях!"
                        )
                    else:
                        # Пользователь в состоянии ожидания файла, но отправил текст
                        send_message(
                            chat_id,
                            "📎 Я жду PDF файл для конвертации.\n\n"
                            "Пожалуйста, отправьте PDF файл или нажмите /start для отмены."
                        )
        elif "document" in message:
            # Проверяем, ожидает ли пользователь загрузки файла
            if not is_user_waiting_for_file(chat_id):
                send_message(
                    chat_id,
                    "📎 Сначала нажмите кнопку «📤 Отправить PDF на конвертацию», а затем отправьте PDF файл.\n\n"
                    "💡 Это поможет мне лучше обработать ваш запрос!"
                )
                return "OK", 200
            
            doc = message["document"]
            if doc.get("mime_type") != "application/pdf":
                send_message(chat_id, "❌ Я принимаю только PDF-файлы. Пожалуйста, отправьте PDF файл.")
                return "OK", 200

            # Проверяем размер файла
            file_size = doc.get("file_size", 0)
            if file_size > 50 * 1024 * 1024:  # 50 МБ
                send_message(chat_id, "❌ Файл слишком большой для обработки. Максимальный размер: 20 МБ")
                return "OK", 200
            
            # Ограничение Telegram Bot API на скачивание файлов напрямую ~20 МБ
            if file_size > 20 * 1024 * 1024:
                send_message(
                    chat_id,
                    "❌ Этот PDF больше 20 МБ. Боты Telegram не могут скачивать такие файлы.\n"
                    "📦 Пожалуйста, сожмите PDF, разбейте на части или пришлите ссылку на файл."
                )
                return "OK", 200

            if file_size > 10 * 1024 * 1024:  # 10 МБ
                send_message(chat_id, "⚠️ Большой файл. Обработка может занять несколько минут...")

            send_message(chat_id, "⏳ Принял PDF. Начинаю обработку...")
            log_event(chat_id, "file_received", {"size": file_size, "name": doc.get("file_name")})

            try:
                file_id = doc["file_id"]
                logger.info(f"📁 Загружаю файл ID: {file_id}")
                
                resp = requests.get(f"https://api.telegram.org/bot{BOT_TOKEN}/getFile?file_id={file_id}", timeout=10)
                if not resp.ok:
                    logger.error(f"❌ Ошибка получения файла: {resp.status_code} - {resp.text}")
                    send_message(chat_id, "❌ Ошибка загрузки файла.")
                    log_event(chat_id, "ocr_error", {"step": "getFile", "status": resp.status_code})
                    log_error(chat_id, "GET_FILE", resp.text)
                    return "OK", 200
                
                file_path = resp.json()["result"]["file_path"]
                file_url = f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path}"
                
                file_resp = requests.get(file_url, timeout=60)  # Увеличили таймаут для больших файлов
                if not file_resp.ok:
                    logger.error(f"❌ Ошибка скачивания файла: {file_resp.status_code}")
                    send_message(chat_id, "❌ Ошибка скачивания файла.")
                    log_event(chat_id, "ocr_error", {"step": "download", "status": file_resp.status_code})
                    log_error(chat_id, "DOWNLOAD_FILE", file_resp.text)
                    return "OK", 200
                
                file_bytes = file_resp.content
                logger.info(f"📄 Файл загружен, размер: {len(file_bytes)} байт")

                # Все PDF файлы обрабатываются через OCR для лучшего качества
                try:
                    reader = PyPDF2.PdfReader(BytesIO(file_bytes))
                    num_pages_detect = len(reader.pages)
                    is_ocr_needed = True  # Всегда используем OCR
                    logger.info(f"🔍 PDF тип: скан (требует OCR) - {num_pages_detect} страниц")
                    log_event(chat_id, "is_ocr", {"pages": num_pages_detect})
                except Exception as e:
                    logger.warning(f"⚠️ Ошибка анализа PDF: {e}")
                    is_ocr_needed = True
                    log_error(chat_id, "ANALYZE_PDF", repr(e))

                # Определим количество страниц (для OCR-сценариев)
                num_pages = 0
                try:
                    reader = PyPDF2.PdfReader(BytesIO(file_bytes))
                    num_pages = len(reader.pages)
                except Exception as e:
                    logger.warning(f"⚠️ Не удалось определить количество страниц: {e}")

                if is_ocr_needed and num_pages > 10:
                    # Сохраняем файл и предлагаем варианты
                    pending_files[chat_id] = {
                        "file_bytes": file_bytes,
                        "file_name": doc.get("file_name", "converted.pdf"),
                        "num_pages": num_pages,
                        "created_at": time.time()
                    }
                    send_message(
                        chat_id,
                        f"🔍 Обнаружен сканированный PDF на {num_pages} страниц.\n\nВыберите, как поступить:",
                        reply_markup=build_split_options_keyboard()
                    )
                    return "OK", 200
                else:
                    if is_ocr_needed:
                        send_message(
                            chat_id,
                            "🔍 Обнаружен скан. Использую OCR. Это займёт 1-3 минуты..."
                    )

                # Функция для отправки прогресса
                def progress_callback(message):
                    logger.info(f"📊 {message}")

                text = extract_text_from_pdf(
                    file_bytes,
                    is_ocr_needed=is_ocr_needed,
                    progress_callback=progress_callback
                )
                if not text.strip():
                    send_message(chat_id, "❌ Не удалось извлечь текст.")
                    log_event(chat_id, "ocr_error", {"file_name": doc.get("file_name")})
                    log_error(chat_id, "OCR_EMPTY", "no text extracted")
                    return "OK", 200

                base_name = doc.get("file_name", "converted")
                txt_name = os.path.splitext(base_name)[0] + ".txt"
                txt_buffer = BytesIO(text.encode("utf-8"))
                send_document(chat_id, txt_buffer, txt_name)
                # Генерируем conversion_id: message_id + timestamp
                conversion_id = f"{message_id}_{int(time.time())}"
                log_event(chat_id, "ocr_success", {"file_name": base_name, "conversion_id": conversion_id})
                send_message(
                    chat_id,
                    "📝 Оцените качество распознавания (1 — плохо, 5 — отлично):",
                    reply_markup=build_rating_keyboard(conversion_id)
                )

                # Сбрасываем состояние ожидания файла
                set_user_waiting_for_file(chat_id, False)

                reply_markup = get_main_keyboard()
                send_message(
                    chat_id,
                    "✅ Готово! Текст успешно извлечён.\n\nНажмите кнопку ниже, чтобы отправить следующий PDF!",
                    reply_markup
                )

            except Exception as e:
                logger.exception("💥 Ошибка при обработке PDF")
                set_user_waiting_for_file(chat_id, False)
                send_message(chat_id, "❌ Произошла ошибка. Попробуйте снова.")
                log_event(chat_id, "ocr_error", {"file_name": message.get('document', {}).get('file_name')})
                log_error(chat_id, "OCR_EXCEPTION", repr(e))

    except Exception as e:
        logger.exception("💥 Критическая ошибка в webhook")
        # Всегда возвращаем 200, чтобы Telegram не повторял запрос
        return "OK", 200

    return "OK", 200

def set_webhook():
    """Устанавливает webhook с улучшенной обработкой ошибок"""
    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/setWebhook"
        webhook_url = WEBHOOK_URL.rstrip("/") + "/webhook"
        resp = requests.post(url, json={"url": webhook_url}, timeout=10)
        if resp.ok:
            logger.info(f"✅ Webhook установлен: {webhook_url}")
        else:
            logger.error(f"❌ Ошибка webhook: {resp.status_code} - {resp.text}")
    except Exception as e:
        logger.exception("💥 Ошибка установки webhook")

if __name__ == "__main__":
    logger.info("🚀 Запуск бота...")
    init_db()
    set_webhook()

    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)